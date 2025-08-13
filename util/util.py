import logging
import os
import sys
import time
from datetime import timedelta, datetime, timezone
from logging.handlers import TimedRotatingFileHandler
from typing import Optional
from pathlib import Path

import pymysql
from dbutils.pooled_db import PooledDB
from contextlib import contextmanager

from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook


class LogUtil:
    _instance = None

    def __new__(cls, *args, **kwargs):
        if not cls._instance:
            cls._instance = super().__new__(cls)
        return cls._instance

    def __init__(
            self,
            name: str = "root",
            log_level: str = "INFO",
            log_dir: str = "logs",
            file_prefix: str = "app",
            file_format: Optional[str] = None
    ):
        if not hasattr(self, '_initialized'):
            self.logger = logging.getLogger(name)
            self.logger.setLevel(log_level)
            os.makedirs(log_dir, exist_ok=True)
            self._init_console_handler()
            self._init_file_handler(log_dir, file_prefix, file_format)
            self._initialized = True

    def _init_console_handler(self) -> None:
        console_handler = logging.StreamHandler(sys.stdout)
        self.logger.addHandler(console_handler)

    def _init_file_handler(
            self,
            log_dir: str,
            prefix: str,
            fmt: Optional[str]
    ) -> None:
        if not fmt:
            fmt = (
                "%(asctime)s [%(levelname)-8s] "
                "%(module)s:%(lineno)d - %(message)s"
            )

        file_handler = TimedRotatingFileHandler(
            filename=os.path.join(log_dir, f"{prefix}.log"),
            when="midnight",
            interval=1,
            backupCount=7,
            encoding="utf-8"
        )

        formatter = logging.Formatter(fmt, datefmt="%Y-%m-%d %H:%M:%S")
        file_handler.setFormatter(formatter)
        self.logger.addHandler(file_handler)

    @classmethod
    def get_logger(cls, name: Optional[str] = None) -> logging.Logger:
        if not cls._instance:
            cls()
        return cls._instance.logger.getChild(name) if name else cls._instance.logger


class SnowFlake:
    def __init__(self, worker_id=1, data_center_id=1):
        self.worker_id = worker_id
        self.data_center_id = data_center_id
        self.sequence = 0
        self.last_timestamp = -1

    def get_timestamp(self):
        return int(time.time() * 1000)

    def wait_for_next_millis(self, last_timestamp):
        timestamp = self.get_timestamp()
        while timestamp <= last_timestamp:
            timestamp = self.get_timestamp()
        return timestamp

    def generate_id(self):
        timestamp = self.get_timestamp()

        if timestamp < self.last_timestamp:
            raise Exception("Clock moved backwards. Refusing to generate id")

        if timestamp == self.last_timestamp:
            self.sequence = (self.sequence + 1) & 0xFFF
            if self.sequence == 0:
                timestamp = self.wait_for_next_millis(self.last_timestamp)
        else:
            self.sequence = 0

        self.last_timestamp = timestamp

        return ((timestamp - 1288834974657) << 22) | (self.data_center_id << 17) | (
                self.worker_id << 12) | self.sequence


class MySQLHelper:
    def __init__(self, db_config: dict,
                 use_pool=True, pool_size=5):

        self.config = db_config

        if use_pool:
            self.pool = PooledDB(
                creator=pymysql,
                maxconnections=pool_size,
                blocking=True,
                **self.config
            )
        else:
            self.pool = None

    @contextmanager
    def get_connection(self):
        conn = None
        try:
            if self.pool:
                conn = self.pool.connection()
            else:
                conn = pymysql.connect(**self.config)
            yield conn
        finally:
            if conn:
                conn.close()

    def execute_query(self, sql, params=None):
        with self.get_connection() as conn:
            with conn.cursor() as cursor:
                cursor.execute(sql, params)
                return cursor.fetchall()

    def execute_update(self, sql, params=None):
        with self.get_connection() as conn:
            with conn.cursor() as cursor:
                affected = cursor.execute(sql, params)
                conn.commit()
                return affected

    def executemany(self, sql, data_list):
        with self.get_connection() as conn:
            with conn.cursor() as cursor:
                affected = cursor.executemany(sql, data_list)
                conn.commit()
                return affected

    def table_exists(self, table_name):
        sql = "SHOW TABLES LIKE %s"
        return bool(self.execute_query(sql, (table_name,)))

    def create_table(self, table_name, columns):
        if self.table_exists(table_name):
            return False

        column_def = ', '.join(columns)
        sql = f"CREATE TABLE {table_name} ({column_def})"
        self.execute_update(sql)
        return True

    @contextmanager
    def transaction(self):
        with self.get_connection() as conn:
            try:
                yield conn
                conn.commit()
            except Exception:
                conn.rollback()
                raise

    def close(self):
        if self.pool:
            self.pool.close()


class CommonUtil:
    def __init__(self):
        self.logger = LogUtil.get_logger("common_util")

    """
        根据时区转换时间戳，并格式化为对应格式
    """

    @staticmethod
    def convert_time_by_zone(time_offset: int, timestamp: int, formatter: str) -> str:
        target_tz = timezone(timedelta(seconds=time_offset))
        utc_time = datetime.fromtimestamp(timestamp / 1000, tz=timezone.utc)
        target_time = utc_time.astimezone(target_tz)
        format_time = target_time.strftime(formatter)
        return format_time

    """
        获取输入文件地址
    """

    @staticmethod
    def get_input_path(input_file_name: str) -> str:
        input_file_dir = Path(__file__).resolve().parent.parent / "input"
        if not input_file_dir.exists():
            input_file_dir.mkdir()
        input_file_path = str(input_file_dir / input_file_name)
        return input_file_path

    """
        获取输出文件地址
    """

    @staticmethod
    def get_output_path(output_file_name: str) -> str:
        output_file_dir = Path(__file__).resolve().parent.parent / "output"
        if not output_file_dir.exists():
            output_file_dir.mkdir()
        output_file_path = output_file_dir / output_file_name
        return str(output_file_path)


class ExcelUtil:
    CLOSE_TYPE_INPUT = "INPUT"
    CLOSE_TYPE_OUTPUT = "OUTPUT"

    def __init__(self, input_file_path: str, output_file_path: str):
        self.logger = LogUtil.get_logger("excel_util")
        if input_file_path is not None and len(input_file_path) > 0 and os.path.exists(input_file_path):
            self.input_file_path = input_file_path
            self.read_wb = load_workbook(input_file_path)
        if output_file_path is not None and len(output_file_path) > 0:
            self.output_file_path = output_file_path
            self.write_wb = Workbook()

    """
        写入excel
        :param data_dict 输入数据，格式：
        {
            "Sheet1": {
                "headers": ['a', 'b', 'c'],
                "data_list": [
                    [1, 2, 3],
                    [4, 5, 6],
                    [7, 8, 9]
                ]
            },
            "Sheet2": {
                "headers": ['a', 'b', 'c'],
                "data_list": [
                    [1, 2, 3],
                    [4, 5, 6],
                    [7, 8, 9]
                ]
            }
        } 
    """

    def write_excel(self, data_dict: dict, append: bool = False):
        for sheet_name, sheet_dict in data_dict.items():
            sheet = self.write_wb.create_sheet(sheet_name)
            headers = sheet_dict["headers"]
            data_list = sheet_dict["data_list"]
            sheet.append(headers)
            for data in data_list:
                sheet.append(data)
        self.write_wb.save(self.output_file_path)
        if not append:
            self.close(type_=self.CLOSE_TYPE_OUTPUT)

    """
        简单读取excel
        读取固定列的数据返回，默认第一行为表头
        输出格式:
        {
            "Sheet1": {
                "Mac": [1,2,3]
            },
            "Sheet2": {
                "add": [4,5,6]
            }
        }
    """

    def simple_read_input_data(self) -> {}:
        data_dict = {}
        sheets = self.read_wb.sheetnames
        for index in range(1, len(sheets)):
            sheet_name = sheets[index - 1]
            sheet_dict = {}
            sheet = self.read_wb[sheet_name]
            headers = [cell.value for cell in sheet[1]]
            for i in range(1, sheet.max_column + 1):
                header = headers[i - 1]
                if header is None:
                    continue
                values = []
                for j in range(2, sheet.max_row + 1):
                    cell = sheet.cell(row=j, column=i)
                    values.append(cell.value)
                sheet_dict[header] = values
            data_dict[sheet_name] = sheet_dict
        return data_dict

    def close(self, type_: str):
        if type_.lower() == "input":
            self.read_wb.close()
        if type_.lower() == "output":
            self.write_wb.save(self.output_file_path)
            self.write_wb.close()
